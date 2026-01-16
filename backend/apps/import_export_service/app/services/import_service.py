import csv
import json
import uuid
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from apps.import_export_service.app.repositories.import_repo import ImportRepository
from apps.import_export_service.app.services.validation_service import ValidationService
from apps.import_export_service.app.services.dedupe_service import DedupeService


class ImportService:
    def __init__(self, db: Session):
        self.import_repo = ImportRepository(db)
        self.validation_service = ValidationService(db)
        self.dedupe_service = DedupeService(db)
        self.db = db
    
    def start_import(
        self,
        user_id: int,
        file_content: bytes,
        file_format: str,
        mode: str = "tolerant"
    ) -> str:
        job_id = str(uuid.uuid4())
        
        self.import_repo.create_job(
            job_id=job_id,
            user_id=user_id,
            format=file_format,
            status="pending"
        )
        
        rows = self._parse_file(file_content, file_format)
        
        errors = []
        imported = 0
        
        for idx, row in enumerate(rows, start=1):
            try:
                validated = self.validation_service.validate_row(row)
                if not self.dedupe_service.is_duplicate(validated):
                    self._import_row(validated, user_id)
                    imported += 1
                else:
                    errors.append({
                        "row_number": idx,
                        "field": "slug",
                        "message": "Duplicate term"
                    })
            except Exception as e:
                errors.append({
                    "row_number": idx,
                    "field": "",
                    "message": str(e)
                })
                
                if mode == "strict":
                    break
        
        self.import_repo.update_job(
            job_id=job_id,
            status="completed",
            imported=imported,
            failed=len(errors)
        )
        
        for error in errors:
            self.import_repo.add_error(
                job_id=job_id,
                row_number=error["row_number"],
                field=error["field"],
                message=error["message"]
            )
        
        return job_id
    
    def _parse_file(self, content: bytes, file_format: str) -> List[Dict[str, Any]]:
        if file_format == "csv":
            return self._parse_csv(content)
        elif file_format == "json":
            return self._parse_json(content)
        elif file_format in ["xlsx", "xls"]:
            return self._parse_excel(content)
        else:
            raise ValueError(f"Unsupported format: {file_format}")
    
    def _parse_csv(self, content: bytes) -> List[Dict[str, Any]]:
        text = content.decode('utf-8')
        reader = csv.DictReader(text.splitlines())
        return list(reader)
    
    def _parse_json(self, content: bytes) -> List[Dict[str, Any]]:
        text = content.decode('utf-8')
        return json.loads(text)
    
    def _parse_excel(self, content: bytes) -> List[Dict[str, Any]]:
        from io import BytesIO
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        rows = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        
        return rows
    
    def _import_row(self, row: Dict[str, Any], user_id: int) -> None:
        pass
